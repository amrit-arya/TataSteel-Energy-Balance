"""
Data Processing Module
======================

Purpose:
    Extract structured data from the raw Excel workbook and transform it
    into clean, normalized DataFrames ready for validation and storage.

Architecture Decision:
    Each gas type is processed by a dedicated extraction function because
    the Excel sheets have different structures:
    - BF Gas: Has both internal (blast furnace) and external consumers
    - CO Gas: Has battery/plant generation and all-external consumers
    - LD Gas: Has generation only — NO consumption data

    The shared gas pool model is enforced here: we do NOT create any
    generator-to-consumer mappings. Each record only links to its gas type.

Data Integrity Rules:
    - Original plant names preserved as-is from the workbook
    - Unique IDs assigned using deterministic patterns
    - No data fabrication: if consumption is unavailable, it stays NULL
    - Units are assumed Nm³/hr (standard for steel plant gas measurement)

Usage:
    python -m analytics.process_data
"""

import os
import sys
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

RAW_DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "raw"
PROCESSED_DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "processed"
WORKBOOK_NAME = "Fuel_Management_Project_Data.xlsx"
WORKBOOK_PATH = RAW_DATA_DIR / WORKBOOK_NAME

# Gas type identifiers
GAS_TYPE_BFG = "BFG"
GAS_TYPE_COG = "COG"
GAS_TYPE_LDG = "LDG"

# Consumer type identifiers
CONSUMER_TYPE_INTERNAL = "Internal"
CONSUMER_TYPE_EXTERNAL = "External"

# Default unit
DEFAULT_UNIT = "Nm³/hr"

# Data unavailable sentinel
DATA_UNAVAILABLE = "Data Unavailable"


# ---------------------------------------------------------------------------
# Data Classes
# ---------------------------------------------------------------------------

@dataclass
class GasType:
    """Represents a gas type in the system."""
    gas_id: str
    gas_name: str
    short_name: str
    description: str = ""


@dataclass
class GenerationSource:
    """Represents a gas generation source."""
    source_id: str
    gas_type_id: str
    source_name: str
    plant_area: str
    generation_value: float
    unit: str = DEFAULT_UNIT


@dataclass
class Consumer:
    """Represents a gas consumer."""
    consumer_id: str
    gas_type_id: str
    consumer_name: str
    consumer_type: str  # 'Internal' or 'External'
    consumption_value: Optional[float] = None  # None if data unavailable
    priority: int = 0
    unit: str = DEFAULT_UNIT


# ---------------------------------------------------------------------------
# ID Generation
# ---------------------------------------------------------------------------

def generate_source_id(gas_type_id: str, index: int) -> str:
    """
    Generate a deterministic unique ID for a generation source.

    Pattern: {GAS_TYPE}-GEN-{NNN}
    Example: BFG-GEN-001, COG-GEN-002
    """
    return f"{gas_type_id}-GEN-{index:03d}"


def generate_consumer_id(gas_type_id: str, consumer_type: str, index: int) -> str:
    """
    Generate a deterministic unique ID for a consumer.

    Pattern: {GAS_TYPE}-CON-{TYPE_CODE}-{NNN}
    Example: BFG-CON-INT-001 (internal), COG-CON-EXT-003 (external)
    """
    type_code = "INT" if consumer_type == CONSUMER_TYPE_INTERNAL else "EXT"
    return f"{gas_type_id}-CON-{type_code}-{index:03d}"


# ---------------------------------------------------------------------------
# Gas Type Definitions
# ---------------------------------------------------------------------------

def create_gas_types() -> list[GasType]:
    """
    Create the three gas type definitions.

    These are derived directly from the workbook sheet names and headers.
    """
    return [
        GasType(
            gas_id=GAS_TYPE_BFG,
            gas_name="Blast Furnace Gas",
            short_name="BF Gas",
            description="By-product gas from blast furnace iron-making process",
        ),
        GasType(
            gas_id=GAS_TYPE_COG,
            gas_name="Coke Oven Gas",
            short_name="CO Gas",
            description="By-product gas from coke oven batteries",
        ),
        GasType(
            gas_id=GAS_TYPE_LDG,
            gas_name="Linz-Donawitz Gas",
            short_name="LD Gas",
            description="By-product gas from LD converter steelmaking process",
        ),
    ]


# ---------------------------------------------------------------------------
# BF Gas Extraction
# ---------------------------------------------------------------------------

def extract_bf_gas(wb: openpyxl.Workbook) -> tuple[list[GenerationSource], list[Consumer]]:
    """
    Extract Blast Furnace Gas data from the 'BF Gas Details' sheet.

    Sheet Structure:
        Row 1:  Title — "BF Gas Overview (Blast Furnace Gas)"
        Row 3:  Section 1 header — "1. Generation & Internal Consumption"
        Row 4:  Column headers — Blast Furnace | Generation | Consumption
        Rows 5-10: Data rows (Furnace I, H, G, F, E, C)
        Row 11: Totals (formula)
        Row 13: Section 2 header — "2. Other Consumptions"
        Row 14: Column headers — Consumer / Plant Area | Consumption
        Rows 15-23: External consumer data rows
        Row 24: Grand Total (formula)
        Row 25: Total Other Consumption (formula)
    """
    ws = wb["BF Gas Details"]
    sources = []
    consumers = []

    # Extract generation sources and internal consumers (rows 5-10)
    gen_index = 1
    int_con_index = 1

    for row_idx in range(5, 11):
        furnace_name = ws.cell(row=row_idx, column=1).value
        generation = ws.cell(row=row_idx, column=2).value
        internal_consumption = ws.cell(row=row_idx, column=3).value

        if furnace_name is None:
            continue

        # Generation source
        sources.append(GenerationSource(
            source_id=generate_source_id(GAS_TYPE_BFG, gen_index),
            gas_type_id=GAS_TYPE_BFG,
            source_name=str(furnace_name).strip(),
            plant_area="Blast Furnace",
            generation_value=float(generation) if generation is not None else 0.0,
        ))
        gen_index += 1

        # Internal consumer (the furnace consumes some of its own gas)
        if internal_consumption is not None:
            consumers.append(Consumer(
                consumer_id=generate_consumer_id(GAS_TYPE_BFG, CONSUMER_TYPE_INTERNAL, int_con_index),
                gas_type_id=GAS_TYPE_BFG,
                consumer_name=str(furnace_name).strip(),
                consumer_type=CONSUMER_TYPE_INTERNAL,
                consumption_value=float(internal_consumption),
                priority=int_con_index,  # Internal consumers get higher priority
            ))
            int_con_index += 1

    # Extract external consumers (rows 15-23)
    ext_con_index = 1

    for row_idx in range(15, 24):
        consumer_name = ws.cell(row=row_idx, column=1).value
        consumption = ws.cell(row=row_idx, column=2).value

        if consumer_name is None:
            continue

        consumers.append(Consumer(
            consumer_id=generate_consumer_id(GAS_TYPE_BFG, CONSUMER_TYPE_EXTERNAL, ext_con_index),
            gas_type_id=GAS_TYPE_BFG,
            consumer_name=str(consumer_name).strip(),
            consumer_type=CONSUMER_TYPE_EXTERNAL,
            consumption_value=float(consumption) if consumption is not None else None,
            priority=10 + ext_con_index,  # External consumers get lower priority
        ))
        ext_con_index += 1

    return sources, consumers


# ---------------------------------------------------------------------------
# CO Gas Extraction
# ---------------------------------------------------------------------------

def extract_co_gas(wb: openpyxl.Workbook) -> tuple[list[GenerationSource], list[Consumer]]:
    """
    Extract Coke Oven Gas data from the 'CO Gas Details' sheet.

    Sheet Structure:
        Row 1:  Title — "CO Gas Overview (Coke Oven Gas)"
        Row 3:  Section 1 header — "1. Generation"
        Row 4:  Column headers — Batteries | Plant / Source | Generation
        Rows 5-6: Generation data (Batt 8,9 / Old BPP; Batt 10,11 / New BPP)
        Row 7:  Total (formula)
        Row 9:  Section 2 header — "2. Consumption"
        Row 10: Column headers — Sl No. | Consumer / Plant Area | Consumption
        Rows 11-28: Consumer data (18 consumers)
        Row 29: Total (formula)
    """
    ws = wb["CO Gas Details"]
    sources = []
    consumers = []

    # Extract generation sources (rows 5-6)
    gen_index = 1

    for row_idx in range(5, 7):
        battery_name = ws.cell(row=row_idx, column=1).value
        plant_source = ws.cell(row=row_idx, column=2).value
        generation = ws.cell(row=row_idx, column=3).value

        if battery_name is None and plant_source is None:
            continue

        # Combine battery and plant name for the source name
        source_name = str(plant_source).strip() if plant_source else str(battery_name).strip()
        plant_area_name = str(battery_name).strip() if battery_name else ""

        sources.append(GenerationSource(
            source_id=generate_source_id(GAS_TYPE_COG, gen_index),
            gas_type_id=GAS_TYPE_COG,
            source_name=source_name,
            plant_area=plant_area_name,
            generation_value=float(generation) if generation is not None else 0.0,
        ))
        gen_index += 1

    # Extract consumers (rows 11-28)
    ext_con_index = 1

    for row_idx in range(11, 29):
        consumer_name = ws.cell(row=row_idx, column=2).value
        consumption = ws.cell(row=row_idx, column=3).value

        if consumer_name is None:
            continue

        consumers.append(Consumer(
            consumer_id=generate_consumer_id(GAS_TYPE_COG, CONSUMER_TYPE_EXTERNAL, ext_con_index),
            gas_type_id=GAS_TYPE_COG,
            consumer_name=str(consumer_name).strip(),
            consumer_type=CONSUMER_TYPE_EXTERNAL,
            consumption_value=float(consumption) if consumption is not None else None,
            priority=ext_con_index,
        ))
        ext_con_index += 1

    return sources, consumers


# ---------------------------------------------------------------------------
# LD Gas Extraction
# ---------------------------------------------------------------------------

def extract_ld_gas(wb: openpyxl.Workbook) -> tuple[list[GenerationSource], list[Consumer]]:
    """
    Extract LD Gas data from the 'LD Gas Details' sheet.

    Sheet Structure:
        Row 1:  Title — "LD Gas Overview (Linz-Donawitz Gas)"
        Row 3:  Section 1 header — "1. Generation"
        Row 4:  Column headers — (empty) | Plant / Source | Generation
        Rows 5-6: Generation data (LD-1 & LD-3; LD-2)
        Row 7:  Total (formula)

    CRITICAL: There is NO consumption section in this sheet.
    The consumers list will be EMPTY — we must NOT fabricate data.
    """
    ws = wb["LD Gas Details"]
    sources = []
    consumers = []  # EMPTY — no consumption data available

    # Extract generation sources (rows 5-6)
    gen_index = 1

    for row_idx in range(5, 7):
        plant_source = ws.cell(row=row_idx, column=2).value
        generation = ws.cell(row=row_idx, column=3).value

        if plant_source is None:
            continue

        sources.append(GenerationSource(
            source_id=generate_source_id(GAS_TYPE_LDG, gen_index),
            gas_type_id=GAS_TYPE_LDG,
            source_name=str(plant_source).strip(),
            plant_area="LD Converter",
            generation_value=float(generation) if generation is not None else 0.0,
        ))
        gen_index += 1

    # NO consumers extracted — consumption data is unavailable
    # This is NOT an error — the workbook genuinely does not contain this data

    return sources, consumers


# ---------------------------------------------------------------------------
# DataFrame Conversion
# ---------------------------------------------------------------------------

def gas_types_to_dataframe(gas_types: list[GasType]) -> pd.DataFrame:
    """Convert gas type objects to DataFrame."""
    return pd.DataFrame([
        {
            "gas_id": gt.gas_id,
            "gas_name": gt.gas_name,
            "short_name": gt.short_name,
            "description": gt.description,
        }
        for gt in gas_types
    ])


def sources_to_dataframe(sources: list[GenerationSource]) -> pd.DataFrame:
    """Convert generation source objects to DataFrame."""
    return pd.DataFrame([
        {
            "source_id": s.source_id,
            "gas_type_id": s.gas_type_id,
            "source_name": s.source_name,
            "plant_area": s.plant_area,
            "generation_value": s.generation_value,
            "unit": s.unit,
        }
        for s in sources
    ])


def consumers_to_dataframe(consumers: list[Consumer]) -> pd.DataFrame:
    """Convert consumer objects to DataFrame."""
    return pd.DataFrame([
        {
            "consumer_id": c.consumer_id,
            "gas_type_id": c.gas_type_id,
            "consumer_name": c.consumer_name,
            "consumer_type": c.consumer_type,
            "consumption_value": c.consumption_value,
            "priority": c.priority,
            "unit": c.unit,
        }
        for c in consumers
    ])


# ---------------------------------------------------------------------------
# Main Processing Pipeline
# ---------------------------------------------------------------------------

def process_workbook() -> dict[str, pd.DataFrame]:
    """
    Run the complete data processing pipeline.

    Returns:
        Dict with DataFrames: gas_types, generation_sources, consumers
    """
    if not WORKBOOK_PATH.exists():
        print(f"ERROR: Workbook not found at {WORKBOOK_PATH}")
        sys.exit(1)

    print(f"Loading workbook: {WORKBOOK_PATH}")
    wb = openpyxl.load_workbook(str(WORKBOOK_PATH), data_only=True)

    # Create gas type definitions
    gas_types = create_gas_types()
    print(f"  Gas types defined: {len(gas_types)}")

    # Extract BF Gas data
    bf_sources, bf_consumers = extract_bf_gas(wb)
    print(f"  BF Gas: {len(bf_sources)} sources, {len(bf_consumers)} consumers "
          f"({sum(1 for c in bf_consumers if c.consumer_type == CONSUMER_TYPE_INTERNAL)} internal, "
          f"{sum(1 for c in bf_consumers if c.consumer_type == CONSUMER_TYPE_EXTERNAL)} external)")

    # Extract CO Gas data
    co_sources, co_consumers = extract_co_gas(wb)
    print(f"  CO Gas: {len(co_sources)} sources, {len(co_consumers)} consumers")

    # Extract LD Gas data
    ld_sources, ld_consumers = extract_ld_gas(wb)
    print(f"  LD Gas: {len(ld_sources)} sources, {len(ld_consumers)} consumers "
          f"(consumption data unavailable)")

    wb.close()

    # Combine all sources and consumers
    all_sources = bf_sources + co_sources + ld_sources
    all_consumers = bf_consumers + co_consumers + ld_consumers

    # Convert to DataFrames
    df_gas_types = gas_types_to_dataframe(gas_types)
    df_sources = sources_to_dataframe(all_sources)
    df_consumers = consumers_to_dataframe(all_consumers)

    # Print summary
    print(f"\n  Total generation sources: {len(df_sources)}")
    print(f"  Total consumers:         {len(df_consumers)}")
    print(f"  Total records:           {len(df_sources) + len(df_consumers)}")

    return {
        "gas_types": df_gas_types,
        "generation_sources": df_sources,
        "consumers": df_consumers,
    }


def save_processed_data(dataframes: dict[str, pd.DataFrame]) -> None:
    """Save processed DataFrames to CSV files."""
    PROCESSED_DATA_DIR.mkdir(parents=True, exist_ok=True)

    for name, df in dataframes.items():
        filepath = PROCESSED_DATA_DIR / f"{name}.csv"
        df.to_csv(filepath, index=False)
        print(f"  Saved: {filepath} ({len(df)} rows)")


# ---------------------------------------------------------------------------
# Entry Point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("=" * 60)
    print("FUEL MANAGEMENT - DATA PROCESSING PIPELINE")
    print("=" * 60)
    print()

    dataframes = process_workbook()

    print()
    print("Saving processed datasets...")
    save_processed_data(dataframes)

    print()
    print("Data processing complete.")
    print("=" * 60)

"""
Workbook Inspection Module
==========================

Purpose:
    Parse and inspect the raw Excel workbook to understand its structure,
    identify data types, detect formulas, and flag any anomalies.

Usage:
    python -m analytics.inspect_workbook

Output:
    Structured inspection report printed to console.
    Summary statistics for each sheet.

Architecture Decision:
    This is the FIRST step in the pipeline. It reads the workbook in two modes:
    - data_only=True  → to see computed values
    - data_only=False → to see formulas
    This dual-pass approach ensures we understand both the raw data
    and the computation logic embedded in the workbook.
"""

import os
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

RAW_DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "raw"
WORKBOOK_NAME = "Fuel_Management_Project_Data.xlsx"
WORKBOOK_PATH = RAW_DATA_DIR / WORKBOOK_NAME


# ---------------------------------------------------------------------------
# Inspection Functions
# ---------------------------------------------------------------------------

def inspect_sheet_structure(ws: Worksheet, sheet_name: str) -> dict:
    """
    Inspect a single worksheet and return its structural metadata.

    Returns a dict containing:
        - dimensions, row/col counts
        - non-empty cell count
        - detected sections/headers
        - data types found
    """
    structure = {
        "sheet_name": sheet_name,
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells": [str(m) for m in ws.merged_cells.ranges],
        "sections": [],
        "data_rows": [],
        "empty_cells_in_data": [],
        "cell_types": set(),
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), start=1):
        row_data = []
        for cell in row:
            if cell.value is not None:
                cell_type = type(cell.value).__name__
                structure["cell_types"].add(cell_type)
                row_data.append({
                    "coordinate": cell.coordinate,
                    "value": cell.value,
                    "type": cell_type,
                })
            else:
                # Track empty cells within data range
                if row_idx > 1:  # Skip title row
                    structure["empty_cells_in_data"].append(cell.coordinate)

        if row_data:
            structure["data_rows"].append(row_data)

    # Convert set to list for serialization
    structure["cell_types"] = list(structure["cell_types"])

    return structure


def inspect_formulas(ws: Worksheet, sheet_name: str) -> list[dict]:
    """
    Inspect a worksheet (loaded without data_only) to find all formulas.

    Returns a list of dicts with formula details.
    """
    formulas = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append({
                    "coordinate": cell.coordinate,
                    "formula": cell.value,
                    "sheet": sheet_name,
                })
    return formulas


def print_inspection_report(
    structures: list[dict],
    all_formulas: list[dict],
    computed_values: dict[str, dict],
) -> None:
    """Print a formatted inspection report to console."""
    print("=" * 70)
    print("FUEL MANAGEMENT WORKBOOK - INSPECTION REPORT")
    print("=" * 70)
    print(f"\nWorkbook: {WORKBOOK_PATH}")
    print(f"Sheets:   {len(structures)}")
    print()

    for structure in structures:
        name = structure["sheet_name"]
        print(f"{'-' * 60}")
        print(f"Sheet: {name}")
        print(f"{'-' * 60}")
        print(f"  Dimensions:    {structure['dimensions']}")
        print(f"  Rows:          {structure['max_row']}")
        print(f"  Columns:       {structure['max_column']}")
        print(f"  Merged Cells:  {structure['merged_cells'] or 'None'}")
        print(f"  Cell Types:    {', '.join(structure['cell_types'])}")
        print(f"  Data Rows:     {len(structure['data_rows'])}")
        print()

        # Print all data rows
        print("  Data Content:")
        for row in structure["data_rows"]:
            cells_str = "  |  ".join(
                f"{c['coordinate']}: {c['value']} ({c['type']})" for c in row
            )
            print(f"    {cells_str}")
        print()

    # Print formulas
    print(f"{'-' * 60}")
    print("FORMULAS DETECTED")
    print(f"{'-' * 60}")
    if all_formulas:
        for f in all_formulas:
            # Get computed value if available
            computed = computed_values.get(f["sheet"], {}).get(f["coordinate"], "N/A")
            print(f"  [{f['sheet']}] {f['coordinate']}: {f['formula']}  ->  {computed}")
    else:
        print("  No formulas detected.")
    print()

    # Summary statistics
    print(f"{'-' * 60}")
    print("SUMMARY STATISTICS")
    print(f"{'-' * 60}")
    total_data_cells = sum(
        sum(1 for _ in row) for s in structures for row in s["data_rows"]
    )
    total_formulas = len(all_formulas)
    print(f"  Total Data Cells:   {total_data_cells}")
    print(f"  Total Formulas:     {total_formulas}")
    print(f"  Total Sheets:       {len(structures)}")
    print()


def get_computed_values(wb_data: openpyxl.Workbook) -> dict[str, dict]:
    """
    Extract computed cell values from data_only workbook.

    Returns nested dict: {sheet_name: {coordinate: value}}
    """
    result = {}
    for sheet_name in wb_data.sheetnames:
        ws = wb_data[sheet_name]
        sheet_values = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            for cell in row:
                if cell.value is not None:
                    sheet_values[cell.coordinate] = cell.value
        result[sheet_name] = sheet_values
    return result


# ---------------------------------------------------------------------------
# Main Entry Point
# ---------------------------------------------------------------------------

def run_inspection() -> dict:
    """
    Run the full workbook inspection.

    Returns:
        dict with keys: structures, formulas, computed_values
    """
    if not WORKBOOK_PATH.exists():
        print(f"ERROR: Workbook not found at {WORKBOOK_PATH}")
        sys.exit(1)

    print(f"Loading workbook: {WORKBOOK_PATH}")

    # Pass 1: Load with computed values
    wb_data = openpyxl.load_workbook(str(WORKBOOK_PATH), data_only=True)
    computed_values = get_computed_values(wb_data)

    # Pass 2: Load with formulas
    wb_formulas = openpyxl.load_workbook(str(WORKBOOK_PATH), data_only=False)

    structures = []
    all_formulas = []

    for sheet_name in wb_data.sheetnames:
        # Structural inspection from data workbook
        ws_data = wb_data[sheet_name]
        structure = inspect_sheet_structure(ws_data, sheet_name)
        structures.append(structure)

        # Formula inspection from formula workbook
        ws_formula = wb_formulas[sheet_name]
        formulas = inspect_formulas(ws_formula, sheet_name)
        all_formulas.extend(formulas)

    # Print report
    print_inspection_report(structures, all_formulas, computed_values)

    wb_data.close()
    wb_formulas.close()

    return {
        "structures": structures,
        "formulas": all_formulas,
        "computed_values": computed_values,
    }


if __name__ == "__main__":
    run_inspection()
